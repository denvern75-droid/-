# -*- coding: utf-8 -*-
"""보고서 파일에서 본문 미리보기 텍스트를 추출함.

선택 의존성(openpyxl / python-docx / pypdf)이 없으면 해당 형식은 건너뛰고
파일 첨부만 진행함 — 의존성 부재로 중계가 멈추지 않게 함.
HWP/HWPX는 hwpx(zip+xml)만 텍스트 추출을 시도함.
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

TAG_RE = re.compile(r"<[^>]+>")
WS_RE = re.compile(r"[ \t ]+")


def _clean(text: str, limit: int) -> str:
    lines = [WS_RE.sub(" ", ln).strip() for ln in text.splitlines()]
    body = "\n".join(ln for ln in lines if ln)
    return body[:limit]


def _from_text(path: Path, limit: int) -> str:
    for enc in ("utf-8-sig", "cp949", "utf-16", "latin-1"):
        try:
            return _clean(path.read_text(encoding=enc), limit)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return ""


def _from_hwpx(path: Path, limit: int) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            names = [n for n in zf.namelist() if n.startswith("Contents/section") and n.endswith(".xml")]
            chunks = []
            for name in sorted(names)[:5]:
                xml = zf.read(name).decode("utf-8", "replace")
                xml = re.sub(r"</hp:p>", "\n", xml)
                chunks.append(TAG_RE.sub("", xml))
                if sum(len(c) for c in chunks) > limit * 4:
                    break
            return _clean("".join(chunks), limit)
    except (zipfile.BadZipFile, KeyError, OSError):
        return ""


def _from_docx(path: Path, limit: int) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", "replace")
        xml = re.sub(r"</w:p>", "\n", xml)
        return _clean(TAG_RE.sub("", xml), limit)
    except (zipfile.BadZipFile, KeyError, OSError):
        return ""


def _from_xlsx(path: Path, limit: int) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return ""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    rows_out: list[str] = []
    try:
        for ws in wb.worksheets[:2]:
            rows_out.append(f"[{ws.title}]")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 30:
                    break
                cells = [str(c) for c in row if c not in (None, "")]
                if cells:
                    rows_out.append(" | ".join(cells))
                if sum(len(r) for r in rows_out) > limit * 2:
                    break
    finally:
        wb.close()
    return _clean("\n".join(rows_out), limit)


def _from_pdf(path: Path, limit: int) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        return ""
    try:
        reader = PdfReader(str(path))
        chunks = [(page.extract_text() or "") for page in reader.pages[:3]]
        return _clean("\n".join(chunks), limit)
    except Exception:
        return ""


_EXTRACTORS = {
    ".txt": _from_text, ".csv": _from_text, ".md": _from_text, ".log": _from_text,
    ".hwpx": _from_hwpx,
    ".docx": _from_docx,
    ".xlsx": _from_xlsx, ".xlsm": _from_xlsx,
    ".pdf": _from_pdf,
}


def extract_preview(path: Path, limit: int = 700) -> str:
    """추출 실패는 예외가 아니라 빈 문자열로 반환함(중계 지속이 우선)."""
    if limit <= 0:
        return ""
    fn = _EXTRACTORS.get(path.suffix.lower())
    if not fn:
        return ""
    try:
        return fn(path, limit)
    except Exception:
        return ""
